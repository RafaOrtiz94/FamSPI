import csv
import json
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

import excel_mapper


class MapperGUI(tk.Tk):
    EXPORT_COLUMNS = [
        "sheet",
        "cell",
        "row",
        "column",
        "source",
        "target_header",
        "value_state",
        "reason",
        "label",
        "description",
    ]

    def __init__(self):
        super().__init__()
        self.title("Mapeador de Sheets - GUI")
        self.geometry("1500x820")
        self.minsize(1200, 680)

        script_dir = Path(__file__).resolve().parent
        self.input_var = tk.StringVar(value="FORMATO BC - 15-01-2026 (2).xlsx")
        self.output_var = tk.StringVar(value=str(script_dir))
        self.format_var = tk.StringVar(value="both")
        self.include_empty_var = tk.BooleanVar(value=False)

        self.view_mode_var = tk.StringVar(value="Objetivo")
        self.sheet_filter_var = tk.StringVar(value="(Todas)")
        self.target_header_filter_var = tk.StringVar(value="(Todas)")
        self.search_var = tk.StringVar(value="")
        self.export_format_var = tk.StringVar(value="csv")

        self.status_var = tk.StringVar(value="Listo para ejecutar.")
        self.summary_var = tk.StringVar(value="Sin resultados todavia.")

        self._rows_objective = []
        self._rows_general = []
        self._rows_all = []
        self._filtered_rows = []
        self._target_header_options = []
        self._selected_target_headers = []

        self._build_ui()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        controls = ttk.LabelFrame(self, text="Configuracion")
        controls.grid(row=0, column=0, sticky="ew", padx=12, pady=10)
        controls.columnconfigure(1, weight=1)

        ttk.Label(controls, text="Archivo Excel:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(controls, textvariable=self.input_var).grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(controls, text="Buscar...", command=self._pick_input).grid(row=0, column=2, padx=8, pady=6)

        ttk.Label(controls, text="Carpeta salida:").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(controls, textvariable=self.output_var).grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(controls, text="Buscar...", command=self._pick_output).grid(row=1, column=2, padx=8, pady=6)

        ttk.Label(controls, text="Formato:").grid(row=2, column=0, sticky="w", padx=8, pady=6)
        ttk.Combobox(
            controls,
            state="readonly",
            values=["json", "md", "both"],
            textvariable=self.format_var,
            width=10,
        ).grid(row=2, column=1, sticky="w", padx=8, pady=6)

        ttk.Checkbutton(
            controls,
            text="Incluir vacias en tabla de celdas detectadas",
            variable=self.include_empty_var,
        ).grid(row=2, column=1, sticky="e", padx=8, pady=6)

        self.run_button = ttk.Button(controls, text="Generar Mapping", command=self._run_clicked)
        self.run_button.grid(row=2, column=2, padx=8, pady=6)

        status_frame = ttk.Frame(self)
        status_frame.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        status_frame.columnconfigure(1, weight=1)

        ttk.Label(status_frame, text="Estado:").grid(row=0, column=0, sticky="w")
        ttk.Label(status_frame, textvariable=self.status_var).grid(row=0, column=1, sticky="w")

        self.progress = ttk.Progressbar(status_frame, mode="indeterminate", length=240)
        self.progress.grid(row=0, column=2, padx=(12, 0), sticky="e")

        ttk.Label(status_frame, textvariable=self.summary_var).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 0))

        result_frame = ttk.LabelFrame(self, text="Busqueda de celdas por rellenar")
        result_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(1, weight=1)

        filter_frame = ttk.Frame(result_frame)
        filter_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=8)
        filter_frame.columnconfigure(4, weight=1)

        ttk.Label(filter_frame, text="Vista:").grid(row=0, column=0, padx=(0, 6), sticky="w")
        self.view_mode_combo = ttk.Combobox(
            filter_frame,
            state="readonly",
            values=["Objetivo", "General", "Ambas"],
            textvariable=self.view_mode_var,
            width=12,
        )
        self.view_mode_combo.grid(row=0, column=1, sticky="w")
        self.view_mode_combo.bind("<<ComboboxSelected>>", self._apply_filter)

        ttk.Label(filter_frame, text="Hoja:").grid(row=0, column=2, padx=(12, 6), sticky="w")
        self.sheet_filter_combo = ttk.Combobox(
            filter_frame,
            state="readonly",
            values=["(Todas)"],
            textvariable=self.sheet_filter_var,
            width=26,
        )
        self.sheet_filter_combo.grid(row=0, column=3, sticky="w")
        self.sheet_filter_combo.bind("<<ComboboxSelected>>", self._apply_filter)

        ttk.Label(filter_frame, text="Columna objetivo:").grid(row=0, column=4, padx=(12, 6), sticky="w")
        self.target_header_filter_entry = ttk.Entry(
            filter_frame,
            textvariable=self.target_header_filter_var,
            state="readonly",
            width=32,
        )
        self.target_header_filter_entry.grid(row=0, column=5, sticky="w")
        ttk.Button(filter_frame, text="Seleccionar...", command=self._select_target_headers).grid(
            row=0,
            column=6,
            padx=(8, 4),
            sticky="w",
        )
        ttk.Button(filter_frame, text="Todas", command=self._clear_target_headers_selection).grid(
            row=0,
            column=7,
            padx=(0, 4),
            sticky="w",
        )

        ttk.Button(filter_frame, text="Limpiar filtros", command=self._clear_filters).grid(row=0, column=8, padx=8)

        ttk.Label(filter_frame, text="Buscar:").grid(row=1, column=0, padx=(0, 6), pady=(8, 0), sticky="w")
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_var)
        self.search_entry.grid(row=1, column=1, columnspan=4, sticky="ew", pady=(8, 0), padx=(0, 8))
        self.search_entry.bind("<KeyRelease>", self._apply_filter)

        ttk.Label(filter_frame, text="Exportar:").grid(row=1, column=5, padx=(4, 6), pady=(8, 0), sticky="e")
        ttk.Combobox(
            filter_frame,
            state="readonly",
            values=["csv", "json", "xlsx"],
            textvariable=self.export_format_var,
            width=8,
        ).grid(row=1, column=6, pady=(8, 0), sticky="w")
        ttk.Button(filter_frame, text="Exportar vista", command=self._export_filtered_view).grid(
            row=1,
            column=7,
            padx=8,
            pady=(8, 0),
            sticky="w",
        )

        columns = (
            "sheet",
            "cell",
            "row",
            "column",
            "source",
            "target_header",
            "value_state",
            "reason",
            "label",
            "description",
        )
        self.targets_tree = ttk.Treeview(result_frame, columns=columns, show="headings")
        self.targets_tree.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))

        self.targets_tree.heading("sheet", text="Hoja")
        self.targets_tree.heading("cell", text="Celda")
        self.targets_tree.heading("row", text="Fila")
        self.targets_tree.heading("column", text="Columna")
        self.targets_tree.heading("source", text="Fuente")
        self.targets_tree.heading("target_header", text="Columna objetivo")
        self.targets_tree.heading("value_state", text="Estado valor")
        self.targets_tree.heading("reason", text="Motivo")
        self.targets_tree.heading("label", text="Contexto/Etiqueta")
        self.targets_tree.heading("description", text="Descripcion de relleno")

        self.targets_tree.column("sheet", width=125, anchor="w")
        self.targets_tree.column("cell", width=80, anchor="center")
        self.targets_tree.column("row", width=70, anchor="center")
        self.targets_tree.column("column", width=95, anchor="center")
        self.targets_tree.column("source", width=120, anchor="center")
        self.targets_tree.column("target_header", width=210, anchor="w")
        self.targets_tree.column("value_state", width=110, anchor="center")
        self.targets_tree.column("reason", width=110, anchor="center")
        self.targets_tree.column("label", width=260, anchor="w")
        self.targets_tree.column("description", width=560, anchor="w")

        y_scroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.targets_tree.yview)
        y_scroll.grid(row=1, column=1, sticky="ns", pady=(0, 8))
        self.targets_tree.configure(yscrollcommand=y_scroll.set)

        x_scroll = ttk.Scrollbar(result_frame, orient="horizontal", command=self.targets_tree.xview)
        x_scroll.grid(row=2, column=0, sticky="ew", padx=8)
        self.targets_tree.configure(xscrollcommand=x_scroll.set)

    def _pick_input(self):
        selected = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(Path(__file__).resolve().parent),
        )
        if selected:
            self.input_var.set(selected)

    def _pick_output(self):
        selected = filedialog.askdirectory(
            title="Seleccionar carpeta de salida",
            initialdir=str(Path(__file__).resolve().parent),
        )
        if selected:
            self.output_var.set(selected)

    def _set_running(self, running):
        if running:
            self.run_button.configure(state="disabled")
            self.progress.start(10)
        else:
            self.run_button.configure(state="normal")
            self.progress.stop()

    def _run_clicked(self):
        input_text = self.input_var.get().strip()
        out_dir_text = self.output_var.get().strip()
        if not input_text:
            messagebox.showerror("Falta archivo", "Debes indicar un archivo Excel .xlsx.")
            return
        if not out_dir_text:
            messagebox.showerror("Falta salida", "Debes indicar una carpeta de salida.")
            return

        self._set_running(True)
        self.status_var.set("Ejecutando mapeador...")
        self.summary_var.set("Procesando archivo, espera unos segundos...")
        self._clear_tree()

        worker = threading.Thread(
            target=self._run_worker,
            args=(input_text, out_dir_text, self.format_var.get(), self.include_empty_var.get()),
            daemon=True,
        )
        worker.start()

    def _run_worker(self, input_text, out_dir_text, output_format, include_empty):
        try:
            input_path = excel_mapper.resolve_input_path(input_text)
            output_dir = Path(out_dir_text).expanduser()

            if not input_path.exists():
                raise FileNotFoundError(f"No se encontro el archivo: {input_path}")
            if input_path.suffix.lower() != ".xlsx":
                raise InvalidFileException("El archivo debe tener extension .xlsx")

            mapping = excel_mapper.analyze_workbook(str(input_path), include_empty=include_empty)
            written_files = excel_mapper.write_outputs(mapping, output_dir, output_format)
            self.after(0, self._on_success, mapping, written_files)
        except Exception as exc:
            self.after(0, self._on_error, exc)

    def _target_to_row(self, sheet_name, target):
        return {
            "sheet": sheet_name,
            "cell": target.get("cell"),
            "row": target.get("row"),
            "column": f"{target.get('column_letter')}/{target.get('column_index')}",
            "source": target.get("source", ""),
            "target_header": target.get("target_header_original") or target.get("target_header") or "",
            "value_state": target.get("value_state", ""),
            "reason": target.get("detection_reason", ""),
            "label": target.get("label") or "(sin contexto)",
            "description": target.get("fill_description", ""),
        }

    def _on_success(self, mapping, written_files):
        self._set_running(False)
        self.status_var.set("Mapping generado correctamente.")

        summary = mapping["summary"]
        self.summary_var.set(
            "Hojas: {0} | Objetivo: {1} | General: {2} | Consolidado: {3}".format(
                summary["total_sheets"],
                summary["empty_fill_targets_objective"],
                summary["empty_fill_targets_general"],
                summary["empty_fill_targets"],
            )
        )

        self._rows_objective = []
        self._rows_general = []
        self._rows_all = []

        for sheet in mapping["sheets"]:
            sheet_name = sheet["name"]
            for target in sheet.get("empty_fill_targets_objective", []):
                self._rows_objective.append(self._target_to_row(sheet_name, target))
            for target in sheet.get("empty_fill_targets_general", []):
                self._rows_general.append(self._target_to_row(sheet_name, target))
            for target in sheet.get("empty_fill_targets", []):
                self._rows_all.append(self._target_to_row(sheet_name, target))

        self._load_filter_values(mapping)
        self._apply_filter()

        output_text = "\n".join(str(file_path.resolve()) for file_path in written_files)
        messagebox.showinfo("Proceso completado", f"Archivos generados:\n{output_text}")

    def _on_error(self, exc):
        self._set_running(False)
        self.status_var.set("Error durante la ejecucion.")
        self.summary_var.set("Revisa el mensaje de error para corregir la configuracion.")
        messagebox.showerror("Error", str(exc))

    def _load_filter_values(self, mapping):
        sheet_values = ["(Todas)"] + [sheet["name"] for sheet in mapping["sheets"]]
        self.sheet_filter_combo.configure(values=sheet_values)
        self.sheet_filter_var.set("(Todas)")

        header_values = set()
        for row in self._rows_objective + self._rows_all:
            if row["target_header"]:
                header_values.add(row["target_header"])

        self._target_header_options = sorted(header_values)
        self._selected_target_headers = [
            header for header in self._selected_target_headers if header in self._target_header_options
        ]
        self._update_target_header_filter_text()

    def _update_target_header_filter_text(self):
        if not self._selected_target_headers:
            self.target_header_filter_var.set("(Todas)")
            return

        if len(self._selected_target_headers) == 1:
            self.target_header_filter_var.set(self._selected_target_headers[0])
            return

        preview = ", ".join(self._selected_target_headers[:2])
        remaining = len(self._selected_target_headers) - 2
        if remaining > 0:
            preview += f" (+{remaining})"
        self.target_header_filter_var.set(preview)

    def _clear_target_headers_selection(self):
        self._selected_target_headers = []
        self._update_target_header_filter_text()
        self._apply_filter()

    def _select_target_headers(self):
        if not self._target_header_options:
            messagebox.showinfo("Sin columnas objetivo", "No hay columnas objetivo detectadas para seleccionar.")
            return

        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar columnas objetivo")
        dialog.geometry("520x420")
        dialog.transient(self)
        dialog.grab_set()

        ttk.Label(
            dialog,
            text="Selecciona una o varias columnas objetivo (Ctrl o Shift para seleccionar multiples).",
            wraplength=480,
        ).pack(anchor="w", padx=12, pady=(12, 6))

        frame = ttk.Frame(dialog)
        frame.pack(fill="both", expand=True, padx=12, pady=6)

        listbox = tk.Listbox(frame, selectmode="extended")
        listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(side="right", fill="y")
        listbox.config(yscrollcommand=scrollbar.set)

        for index, header in enumerate(self._target_header_options):
            listbox.insert("end", header)
            if header in self._selected_target_headers:
                listbox.selection_set(index)

        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill="x", padx=12, pady=(0, 12))

        def apply_selection():
            selected_indices = listbox.curselection()
            self._selected_target_headers = [self._target_header_options[i] for i in selected_indices]
            self._update_target_header_filter_text()
            self._apply_filter()
            dialog.destroy()

        ttk.Button(button_frame, text="Aplicar", command=apply_selection).pack(side="right")
        ttk.Button(button_frame, text="Cancelar", command=dialog.destroy).pack(side="right", padx=(0, 8))

    def _clear_filters(self):
        self.view_mode_var.set("Objetivo")
        self.sheet_filter_var.set("(Todas)")
        self._selected_target_headers = []
        self._update_target_header_filter_text()
        self.search_var.set("")
        self._apply_filter()

    def _get_rows_by_view_mode(self):
        view_mode = self.view_mode_var.get()
        if view_mode == "General":
            return self._rows_general
        if view_mode == "Ambas":
            return self._rows_all
        return self._rows_objective

    def _apply_filter(self, _event=None):
        rows = list(self._get_rows_by_view_mode())

        selected_sheet = self.sheet_filter_var.get().strip()
        if selected_sheet and selected_sheet != "(Todas)":
            rows = [row for row in rows if row["sheet"] == selected_sheet]

        if self._selected_target_headers:
            selected_set = set(self._selected_target_headers)
            rows = [row for row in rows if row["target_header"] in selected_set]

        search_text = self.search_var.get().strip().lower()
        if search_text:

            def matches_search(row):
                haystack = " ".join(
                    [
                        str(row["sheet"]),
                        str(row["cell"]),
                        str(row["column"]),
                        str(row["source"]),
                        str(row["target_header"]),
                        str(row["value_state"]),
                        str(row["reason"]),
                        str(row["label"]),
                        str(row["description"]),
                    ]
                ).lower()
                return search_text in haystack

            rows = [row for row in rows if matches_search(row)]

        self._filtered_rows = rows

        self._clear_tree()
        for row in rows:
            self.targets_tree.insert(
                "",
                "end",
                values=(
                    row["sheet"],
                    row["cell"],
                    row["row"],
                    row["column"],
                    row["source"],
                    row["target_header"],
                    row["value_state"],
                    row["reason"],
                    row["label"],
                    row["description"],
                ),
            )

    def _default_export_name(self, export_format):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        view_mode = self.view_mode_var.get().strip().lower().replace(" ", "_")
        return f"mapping_filtrado_{view_mode}_{timestamp}.{export_format}"

    def _export_filtered_view(self):
        rows = list(self._filtered_rows)
        if not rows:
            messagebox.showwarning("Sin datos", "No hay filas en la vista filtrada para exportar.")
            return

        export_format = self.export_format_var.get().strip().lower()
        if export_format not in {"csv", "json", "xlsx"}:
            messagebox.showerror("Formato invalido", f"Formato de exportacion no soportado: {export_format}")
            return

        initial_dir = Path(self.output_var.get().strip() or Path(__file__).resolve().parent)
        default_name = self._default_export_name(export_format)

        save_path = filedialog.asksaveasfilename(
            title="Exportar vista filtrada",
            initialdir=str(initial_dir),
            initialfile=default_name,
            defaultextension=f".{export_format}",
            filetypes=[
                ("CSV", "*.csv"),
                ("JSON", "*.json"),
                ("Excel", "*.xlsx"),
            ],
        )

        if not save_path:
            return

        output_path = Path(save_path)

        try:
            if export_format == "csv":
                self._write_rows_csv(output_path, rows)
            elif export_format == "json":
                self._write_rows_json(output_path, rows)
            else:
                self._write_rows_xlsx(output_path, rows)

            messagebox.showinfo("Exportacion completada", f"Archivo exportado:\n{output_path.resolve()}")
        except Exception as exc:
            messagebox.showerror("Error exportando", str(exc))

    def _write_rows_csv(self, path, rows):
        with open(path, "w", encoding="utf-8", newline="") as file_obj:
            writer = csv.DictWriter(file_obj, fieldnames=self.EXPORT_COLUMNS)
            writer.writeheader()
            for row in rows:
                writer.writerow({column: row.get(column, "") for column in self.EXPORT_COLUMNS})

    def _write_rows_json(self, path, rows):
        payload = {
            "exported_at": datetime.now().isoformat(),
            "view_mode": self.view_mode_var.get(),
            "sheet_filter": self.sheet_filter_var.get(),
            "target_header_filter": list(self._selected_target_headers),
            "search": self.search_var.get(),
            "rows": rows,
        }
        with open(path, "w", encoding="utf-8") as file_obj:
            json.dump(payload, file_obj, indent=2, ensure_ascii=False)

    def _write_rows_xlsx(self, path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "mapping_filtrado"

        for col_idx, name in enumerate(self.EXPORT_COLUMNS, start=1):
            sheet.cell(row=1, column=col_idx, value=name)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, name in enumerate(self.EXPORT_COLUMNS, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row.get(name, ""))

        workbook.save(path)

    def _clear_tree(self):
        for item in self.targets_tree.get_children():
            self.targets_tree.delete(item)


def main():
    app = MapperGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
