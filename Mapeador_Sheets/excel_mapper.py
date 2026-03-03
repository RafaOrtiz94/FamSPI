import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

TARGET_HEADER_DET = "DET/AÑO PROCESO"
TARGET_HEADER_PRODUCT = "PRODUCTO A ENTREGAR"


def detect_data_validations(sheet):
    validations = {}
    for dv in sheet.data_validations.dataValidation:
        for cell_range in dv.ranges:
            for row in sheet.iter_rows(
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
            ):
                for cell in row:
                    validations[cell.coordinate] = {
                        "type": dv.type,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                    }
    return validations


def build_merged_cell_helpers(sheet):
    merged_non_anchor_cells = set()
    merged_anchor_cells = {}

    for merged_range in sheet.merged_cells.ranges:
        anchor = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        merged_anchor_cells[anchor] = str(merged_range)

        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                if coord != anchor:
                    merged_non_anchor_cells.add(coord)

    return merged_non_anchor_cells, merged_anchor_cells


def get_used_bounds(sheet, validations):
    min_col, min_row, max_col, max_row = range_boundaries(sheet.calculate_dimension())

    for coord in validations:
        row_idx, col_idx = coordinate_to_tuple(coord)
        min_col = min(min_col, col_idx)
        min_row = min(min_row, row_idx)
        max_col = max(max_col, col_idx)
        max_row = max(max_row, row_idx)

    for merged in sheet.merged_cells.ranges:
        min_col = min(min_col, merged.min_col)
        min_row = min(min_row, merged.min_row)
        max_col = max(max_col, merged.max_col)
        max_row = max(max_row, merged.max_row)

    return {
        "min_col": min_col,
        "min_row": min_row,
        "max_col": max_col,
        "max_row": max_row,
    }


def normalize_text(value):
    text = str(value).strip().upper()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.replace("/", " ")
    normalized = re.sub(r"[^A-Z0-9 ]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def classify_fillable_header(value):
    normalized = normalize_text(value)
    if not normalized:
        return None

    tokens = set(normalized.split())

    if "PRODUCTO" in tokens and "ENTREGAR" in tokens:
        return TARGET_HEADER_PRODUCT

    has_process_year = "PROCESO" in tokens and "ANO" in tokens
    has_det_or_quantity = "DET" in tokens or "CANTIDAD" in tokens
    if has_process_year and has_det_or_quantity:
        return TARGET_HEADER_DET

    return None


def is_cell_empty(cell):
    if cell.data_type == "f":
        return False
    if cell.value is None:
        return True
    if isinstance(cell.value, str) and not cell.value.strip():
        return True
    return False


def extract_label_text(cell):
    if cell.data_type == "f":
        return None

    value = cell.value
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        return cleaned.replace("\n", " ")

    return str(value)


def find_nearest_label(sheet, row_idx, col_idx, max_left_steps=2, max_up_steps=1):
    for step in range(1, max_left_steps + 1):
        left_col = col_idx - step
        if left_col >= 1:
            candidate_cell = sheet.cell(row=row_idx, column=left_col)
            label_text = extract_label_text(candidate_cell)
            if label_text:
                return {
                    "label": label_text,
                    "label_cell": candidate_cell.coordinate,
                    "direction": "left",
                }

    for step in range(1, max_up_steps + 1):
        upper_row = row_idx - step
        if upper_row >= 1:
            candidate_cell = sheet.cell(row=upper_row, column=col_idx)
            label_text = extract_label_text(candidate_cell)
            if label_text:
                return {
                    "label": label_text,
                    "label_cell": candidate_cell.coordinate,
                    "direction": "up",
                }

    return None


def build_general_fill_description(sheet_name, cell_coord, row_idx, col_idx, label_context, has_validation):
    column_letter = get_column_letter(col_idx)

    if label_context:
        return (
            f"Rellenar '{label_context['label']}' en {sheet_name}!{cell_coord} "
            f"(fila {row_idx}, columna {column_letter}/{col_idx}). "
            f"Referencia detectada en {label_context['label_cell']} ({label_context['direction']})."
        )

    if has_validation:
        return (
            f"Rellenar celda con validacion en {sheet_name}!{cell_coord} "
            f"(fila {row_idx}, columna {column_letter}/{col_idx})."
        )

    return (
        f"Rellenar celda vacia en {sheet_name}!{cell_coord} "
        f"(fila {row_idx}, columna {column_letter}/{col_idx})."
    )


def build_general_fill_target(sheet, cell, validations):
    row_idx = cell.row
    col_idx = cell.column
    label_context = find_nearest_label(sheet, row_idx, col_idx)
    has_validation = cell.coordinate in validations

    if not (has_validation or label_context):
        return None

    detection_reason = "validation" if has_validation else "label_context"

    return {
        "cell": cell.coordinate,
        "row": row_idx,
        "column_letter": get_column_letter(col_idx),
        "column_index": col_idx,
        "source": "general",
        "detection_reason": detection_reason,
        "target_header": None,
        "target_header_cell": None,
        "target_header_original": None,
        "has_validation": has_validation,
        "validation": validations.get(cell.coordinate),
        "label": label_context["label"] if label_context else None,
        "label_cell": label_context["label_cell"] if label_context else None,
        "label_direction": label_context["direction"] if label_context else None,
        "fill_description": build_general_fill_description(
            sheet.title,
            cell.coordinate,
            row_idx,
            col_idx,
            label_context,
            has_validation,
        ),
    }


def detect_fillable_headers(sheet, merged_non_anchor_cells, bounds):
    headers = []

    for row in sheet.iter_rows(
        min_row=bounds["min_row"],
        max_row=bounds["max_row"],
        min_col=bounds["min_col"],
        max_col=bounds["max_col"],
    ):
        for cell in row:
            if cell.coordinate in merged_non_anchor_cells:
                continue
            if not isinstance(cell.value, str):
                continue

            target_header = classify_fillable_header(cell.value)
            if not target_header:
                continue

            headers.append(
                {
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "target_header": target_header,
                    "header_text": cell.value.strip(),
                }
            )

    headers.sort(key=lambda item: (item["row"], item["column"]))
    return headers


def find_row_context_for_target(sheet, row_idx, target_col):
    string_fallback = None
    any_fallback = None

    for col_idx in range(target_col - 1, 0, -1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        label_text = extract_label_text(cell)
        if not label_text:
            continue

        context = {
            "label": label_text,
            "cell": cell.coordinate,
        }

        if isinstance(cell.value, str):
            return context

        if not string_fallback:
            string_fallback = context
        if not any_fallback:
            any_fallback = context

    return string_fallback or any_fallback


def build_target_fill_description(sheet_name, target_cell, header_info, row_context):
    row_idx = target_cell.row
    col_idx = target_cell.column
    col_letter = get_column_letter(col_idx)
    header_text = header_info["header_text"]

    description = (
        f"Rellenar columna '{header_text}' en {sheet_name}!{target_cell.coordinate} "
        f"(fila {row_idx}, columna {col_letter}/{col_idx})."
    )

    if row_context:
        description += (
            f" Contexto de fila: '{row_context['label']}' en {row_context['cell']}."
        )

    return description


def build_fill_targets_by_headers(sheet, headers, validations):
    fill_targets = []
    if not headers:
        return fill_targets

    header_rows = [header["row"] for header in headers]

    for header in headers:
        header_row = header["row"]
        target_col = header["column"]
        next_rows = [row_value for row_value in header_rows if row_value > header_row]
        max_row = min(next_rows) - 1 if next_rows else sheet.max_row

        for row_idx in range(header_row + 1, max_row + 1):
            target_cell = sheet.cell(row=row_idx, column=target_col)
            row_context = find_row_context_for_target(sheet, row_idx, target_col)

            has_target_value = not is_cell_empty(target_cell)
            has_context = row_context is not None

            if not has_target_value and not has_context:
                continue

            if has_target_value:
                continue

            fill_targets.append(
                {
                    "cell": target_cell.coordinate,
                    "row": row_idx,
                    "column_letter": get_column_letter(target_col),
                    "column_index": target_col,
                    "source": "objective",
                    "detection_reason": "target_column",
                    "target_header": header["target_header"],
                    "target_header_cell": header["cell"],
                    "target_header_original": header["header_text"],
                    "has_validation": target_cell.coordinate in validations,
                    "validation": validations.get(target_cell.coordinate),
                    "label": row_context["label"] if row_context else None,
                    "label_cell": row_context["cell"] if row_context else None,
                    "label_direction": "left_row_context" if row_context else None,
                    "fill_description": build_target_fill_description(
                        sheet.title,
                        target_cell,
                        header,
                        row_context,
                    ),
                }
            )

    return fill_targets


def collect_general_fill_targets(sheet, validations, bounds, merged_non_anchor_cells):
    targets = []
    for row in sheet.iter_rows(
        min_row=bounds["min_row"],
        max_row=bounds["max_row"],
        min_col=bounds["min_col"],
        max_col=bounds["max_col"],
    ):
        for cell in row:
            if cell.coordinate in merged_non_anchor_cells:
                continue
            if not is_cell_empty(cell):
                continue

            target = build_general_fill_target(sheet, cell, validations)
            if target:
                targets.append(target)

    return targets


def merge_fill_targets(objective_targets, general_targets):
    merged = {}

    for target in objective_targets:
        merged[target["cell"]] = {**target}

    for target in general_targets:
        cell = target["cell"]
        if cell in merged:
            existing = merged[cell]
            existing["source"] = "objective+general"
            existing.setdefault("extra_detection_reasons", [])
            if target["detection_reason"] not in existing["extra_detection_reasons"]:
                existing["extra_detection_reasons"].append(target["detection_reason"])
        else:
            merged[cell] = {**target}

    return sorted(merged.values(), key=lambda item: (item["row"], item["column_index"]))


def analyze_sheet(sheet, include_empty=False):
    sheet_info = {
        "name": sheet.title,
        "merged_cells": [],
        "cells": [],
        "fillable_headers": [],
        "empty_fill_targets_objective": [],
        "empty_fill_targets_general": [],
        "empty_fill_targets": [],
        "summary": {},
    }

    for merged in sheet.merged_cells.ranges:
        sheet_info["merged_cells"].append(str(merged))

    validations = detect_data_validations(sheet)
    bounds = get_used_bounds(sheet, validations)
    merged_non_anchor_cells, _ = build_merged_cell_helpers(sheet)

    fillable_headers = detect_fillable_headers(sheet, merged_non_anchor_cells, bounds)
    objective_targets = build_fill_targets_by_headers(sheet, fillable_headers, validations)
    general_targets = collect_general_fill_targets(sheet, validations, bounds, merged_non_anchor_cells)
    merged_targets = merge_fill_targets(objective_targets, general_targets)

    sheet_info["fillable_headers"] = fillable_headers
    sheet_info["empty_fill_targets_objective"] = objective_targets
    sheet_info["empty_fill_targets_general"] = general_targets
    sheet_info["empty_fill_targets"] = merged_targets

    formulas_count = 0
    empty_cells_in_range = 0

    for row in sheet.iter_rows(
        min_row=bounds["min_row"],
        max_row=bounds["max_row"],
        min_col=bounds["min_col"],
        max_col=bounds["max_col"],
    ):
        for cell in row:
            if cell.coordinate in merged_non_anchor_cells:
                continue

            is_formula = cell.data_type == "f"
            empty_cell = is_cell_empty(cell)

            if empty_cell:
                empty_cells_in_range += 1

            if empty_cell and not include_empty:
                continue

            has_validation = cell.coordinate in validations
            if is_formula:
                formulas_count += 1

            cell_data = {
                "cell": cell.coordinate,
                "value": cell.value,
                "data_type": cell.data_type,
                "is_formula": is_formula,
                "has_validation": has_validation,
                "validation": validations.get(cell.coordinate),
                "row": cell.row,
                "column_letter": get_column_letter(cell.column),
                "column_index": cell.column,
                "is_empty": empty_cell,
            }
            sheet_info["cells"].append(cell_data)

    sheet_info["summary"] = {
        "mapped_cells": len(sheet_info["cells"]),
        "formula_cells": formulas_count,
        "validation_cells": len(validations),
        "merged_ranges": len(sheet_info["merged_cells"]),
        "fillable_headers": len(fillable_headers),
        "empty_cells_in_range": empty_cells_in_range,
        "empty_fill_targets_objective": len(objective_targets),
        "empty_fill_targets_general": len(general_targets),
        "empty_fill_targets": len(merged_targets),
    }
    return sheet_info


def analyze_workbook(path, include_empty=False):
    workbook = load_workbook(path, data_only=False)

    workbook_info = {
        "file": str(path),
        "sheets": [],
        "summary": {
            "total_sheets": len(workbook.sheetnames),
            "mapped_cells": 0,
            "formula_cells": 0,
            "validation_cells": 0,
            "merged_ranges": 0,
            "fillable_headers": 0,
            "empty_cells_in_range": 0,
            "empty_fill_targets_objective": 0,
            "empty_fill_targets_general": 0,
            "empty_fill_targets": 0,
        },
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_info = analyze_sheet(sheet, include_empty=include_empty)
        workbook_info["sheets"].append(sheet_info)

        workbook_info["summary"]["mapped_cells"] += sheet_info["summary"]["mapped_cells"]
        workbook_info["summary"]["formula_cells"] += sheet_info["summary"]["formula_cells"]
        workbook_info["summary"]["validation_cells"] += sheet_info["summary"]["validation_cells"]
        workbook_info["summary"]["merged_ranges"] += sheet_info["summary"]["merged_ranges"]
        workbook_info["summary"]["fillable_headers"] += sheet_info["summary"]["fillable_headers"]
        workbook_info["summary"]["empty_cells_in_range"] += sheet_info["summary"]["empty_cells_in_range"]
        workbook_info["summary"]["empty_fill_targets_objective"] += sheet_info["summary"]["empty_fill_targets_objective"]
        workbook_info["summary"]["empty_fill_targets_general"] += sheet_info["summary"]["empty_fill_targets_general"]
        workbook_info["summary"]["empty_fill_targets"] += sheet_info["summary"]["empty_fill_targets"]

    return workbook_info


def _append_targets_table(markdown, title, targets, include_target_header=False):
    markdown += f"### {title}\n\n"
    if not targets:
        markdown += "No se detectaron celdas para esta vista.\n\n"
        return markdown

    if include_target_header:
        markdown += "| Celda | Fila | Columna | Columna objetivo | Motivo | Contexto fila | Descripcion de relleno |\n"
        markdown += "|-------|------|---------|------------------|--------|---------------|--------------------------|\n"
        for target in targets:
            context_label = target["label"] if target["label"] else "(sin contexto)"
            description = target["fill_description"].replace("|", "\\|")
            target_header = target.get("target_header_original") or target.get("target_header")
            markdown += (
                f"| {target['cell']} | {target['row']} | {target['column_letter']}/{target['column_index']} "
                f"| {target_header} | {target['detection_reason']} | {context_label} | {description} |\n"
            )
    else:
        markdown += "| Celda | Fila | Columna | Motivo | Etiqueta detectada | Descripcion de relleno |\n"
        markdown += "|-------|------|---------|--------|--------------------|--------------------------|\n"
        for target in targets:
            label = target["label"] if target["label"] else "(sin etiqueta cercana)"
            description = target["fill_description"].replace("|", "\\|")
            markdown += (
                f"| {target['cell']} | {target['row']} | {target['column_letter']}/{target['column_index']} "
                f"| {target['detection_reason']} | {label} | {description} |\n"
            )

    markdown += "\n"
    return markdown


def generate_markdown(mapping):
    markdown = "# Excel Mapping Report\n\n"
    markdown += f"Archivo: {mapping['file']}\n\n"
    markdown += "## Resumen General\n\n"
    markdown += f"- Hojas: {mapping['summary']['total_sheets']}\n"
    markdown += f"- Celdas mapeadas: {mapping['summary']['mapped_cells']}\n"
    markdown += f"- Celdas con formula: {mapping['summary']['formula_cells']}\n"
    markdown += f"- Celdas con validacion: {mapping['summary']['validation_cells']}\n"
    markdown += f"- Rangos combinados: {mapping['summary']['merged_ranges']}\n"
    markdown += f"- Encabezados objetivo detectados: {mapping['summary']['fillable_headers']}\n"
    markdown += f"- Celdas vacias en rango usado: {mapping['summary']['empty_cells_in_range']}\n"
    markdown += f"- Celdas vacias por rellenar (objetivo): {mapping['summary']['empty_fill_targets_objective']}\n"
    markdown += f"- Celdas vacias por rellenar (general): {mapping['summary']['empty_fill_targets_general']}\n"
    markdown += f"- Celdas vacias por rellenar (consolidado): {mapping['summary']['empty_fill_targets']}\n\n"

    for sheet in mapping["sheets"]:
        markdown += f"## Hoja: {sheet['name']}\n\n"
        markdown += "### Resumen de Hoja\n"
        markdown += f"- Celdas mapeadas: {sheet['summary']['mapped_cells']}\n"
        markdown += f"- Celdas con formula: {sheet['summary']['formula_cells']}\n"
        markdown += f"- Celdas con validacion: {sheet['summary']['validation_cells']}\n"
        markdown += f"- Rangos combinados: {sheet['summary']['merged_ranges']}\n"
        markdown += f"- Encabezados objetivo detectados: {sheet['summary']['fillable_headers']}\n"
        markdown += f"- Celdas vacias en rango usado: {sheet['summary']['empty_cells_in_range']}\n"
        markdown += f"- Celdas vacias por rellenar (objetivo): {sheet['summary']['empty_fill_targets_objective']}\n"
        markdown += f"- Celdas vacias por rellenar (general): {sheet['summary']['empty_fill_targets_general']}\n"
        markdown += f"- Celdas vacias por rellenar (consolidado): {sheet['summary']['empty_fill_targets']}\n\n"

        markdown = _append_targets_table(
            markdown,
            "Celdas Vacias por Rellenar (Columnas Objetivo)",
            sheet["empty_fill_targets_objective"],
            include_target_header=True,
        )

        markdown = _append_targets_table(
            markdown,
            "Celdas Vacias por Rellenar (General - Informe Completo)",
            sheet["empty_fill_targets_general"],
            include_target_header=False,
        )

        markdown = _append_targets_table(
            markdown,
            "Celdas Vacias por Rellenar (Consolidado)",
            sheet["empty_fill_targets"],
            include_target_header=True,
        )

        if sheet["merged_cells"]:
            markdown += "### Celdas Combinadas\n"
            for merged in sheet["merged_cells"]:
                markdown += f"- {merged}\n"
            markdown += "\n"

        markdown += "### Celdas Detectadas\n\n"
        markdown += "| Celda | Fila | Columna | Valor | Formula | Validacion | Vacia |\n"
        markdown += "|-------|------|---------|-------|---------|------------|-------|\n"
        for cell in sheet["cells"]:
            value = str(cell["value"]).replace("\n", " ").replace("|", "\\|")
            markdown += (
                f"| {cell['cell']} | {cell['row']} | {cell['column_letter']}/{cell['column_index']} "
                f"| {value} | {cell['is_formula']} | {cell['has_validation']} | {cell['is_empty']} |\n"
            )
        markdown += "\n\n"

    return markdown


def resolve_input_path(file_name):
    script_dir = Path(__file__).resolve().parent
    file_path = Path(file_name)
    if not file_path.is_absolute() and not file_path.exists():
        file_path = script_dir / file_name
    return file_path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Mapea estructura y contenido de archivos Excel.")
    parser.add_argument(
        "--input",
        default="FORMATO BC - 15-01-2026 (2).xlsx",
        help="Archivo .xlsx a analizar. Si es relativo, primero intenta en el directorio actual y luego junto al script.",
    )
    parser.add_argument(
        "--out-dir",
        default=str(Path(__file__).resolve().parent),
        help="Directorio de salida para los reportes.",
    )
    parser.add_argument(
        "--format",
        choices=("json", "md", "both"),
        default="both",
        help="Formato de salida.",
    )
    parser.add_argument(
        "--include-empty",
        action="store_true",
        help="Incluye celdas vacias en la tabla de celdas detectadas.",
    )
    return parser.parse_args(argv)


def write_outputs(mapping, out_dir, output_format):
    out_dir.mkdir(parents=True, exist_ok=True)
    written_files = []

    if output_format in ("json", "both"):
        json_output = out_dir / "mapping_auto.json"
        with open(json_output, "w", encoding="utf-8") as file_obj:
            json.dump(mapping, file_obj, indent=2, default=str)
        written_files.append(json_output)

    if output_format in ("md", "both"):
        md_output = out_dir / "mapping_auto.md"
        with open(md_output, "w", encoding="utf-8") as file_obj:
            file_obj.write(generate_markdown(mapping))
        written_files.append(md_output)

    return written_files


def main(argv=None):
    args = parse_args(argv)
    input_path = resolve_input_path(args.input)
    out_dir = Path(args.out_dir)

    try:
        if not input_path.exists():
            raise FileNotFoundError(f"No se encontro el archivo: {input_path}")
        if input_path.suffix.lower() != ".xlsx":
            raise InvalidFileException("El archivo debe tener extension .xlsx")

        mapping = analyze_workbook(str(input_path), include_empty=args.include_empty)
        output_files = write_outputs(mapping, out_dir, args.format)

        print("[OK] Mapping generado:")
        for output_path in output_files:
            print(f" - {output_path.resolve()}")
        return 0
    except FileNotFoundError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    except InvalidFileException as exc:
        print(f"[ERROR] Archivo Excel invalido: {exc}", file=sys.stderr)
        return 2
    except PermissionError as exc:
        print(f"[ERROR] Sin permisos para leer/escribir archivos: {exc}", file=sys.stderr)
        return 3
    except Exception as exc:
        print(f"[ERROR] Fallo inesperado: {exc}", file=sys.stderr)
        return 99


if __name__ == "__main__":
    raise SystemExit(main())
